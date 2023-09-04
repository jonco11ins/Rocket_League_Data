import pandas as pd
import os
import openpyxl
from openpyxl import load_workbook
import glob
from openpyxl.styles import Font
from datetime import datetime


# Moving files from downloads to players_analytics folder
source_folder = '/Users/jonco11ins/downloads'
destination_folder = '/Users/jonco11ins/documents/players_analytics'
# List all files in the source folder
files = os.listdir(source_folder)
# Iterate through the files and move those that match the pattern
#this is inplace incase I were to manually duplicate or move a file
for file in files:
    if file.endswith('players.csv'):
        source_file = os.path.join(source_folder, file)
        destination_file = os.path.join(destination_folder, file)
        # Move the file using os.rename
        os.rename(source_file, destination_file)
        print(f"Moved '{file}' to '{destination_folder}'")
# Using glob to find files matching the pattern 'players (1)' in the destination folder
pattern = os.path.join(destination_folder, '*players (1)*')
matching_files = glob.glob(pattern)
# Iterate through the matching files and delete them
for file in matching_files:
    os.remove(file)  
    print(f"Deleted duplicate'{file}'")


#INDICATES GAMES THAT DONT HAVE 'HIHATER1292' or 'JC CUTS IN FILE, THIS CAN BE REOMVED WITH FILTERS BUT I WANT IT TO ONLY CONTAIN GAMES WE PLAYED
directory_path = '/Users/jonco11ins/Documents/players_analytics'
# List to store file names without "hihater1297"
files_to_delete = []
# Iterate through all files in the directory
for file in os.listdir(directory_path):
    if file.endswith('players.csv'):
        with open(os.path.join(directory_path, file), 'r') as csv_file:
            content = csv_file.read()
            if "hihater1297" and "JC CUTS" not in content:
                files_to_delete.append(file)
# Print the filtered file names
for file in files_to_delete:
    os.remove(os.path.join(directory_path,file))
    print(f"Deleted non-relevant'{file}'")    


#ADD FIELDS TO PLAYERS DATA: Function to calculate 'Teams Goals' and 'Outcome' for each set of 4 rows (match)
def update_teams_goals_and_outcome(df):
    for i in range(0, len(df), 4):
        teams_goals1 = df.loc[i:i + 1, 'goals'].sum()
        teams_goals2 = df.loc[i + 2:i + 3, 'goals'].sum()
        df.loc[i:i + 1, 'Outcome'] = ['W' if teams_goals1 > teams_goals2 else 'L' if teams_goals1 < teams_goals2 else 'D'] * 2
        df.loc[i + 2:i + 3, 'Outcome'] = ['L' if teams_goals1 > teams_goals2 else 'W' if teams_goals1 < teams_goals2 else 'D'] * 2
        df.loc[i:i + 1, 'Teams Goals'] = [teams_goals1] * 2
        df.loc[i + 2:i + 3, 'Teams Goals'] = [teams_goals2] * 2
# Define the directory path
directory_paths = '/Users/jonco11ins/documents/players_analytics'
# List all CSV files in the directory
csv_files = [file for file in os.listdir(directory_paths) if file.endswith('-players.csv')]
for csv_file in csv_files:
    file_path = os.path.join(directory_paths, csv_file)
    # Read the CSV file
    df = pd.read_csv(file_path, sep=';')
    # Apply the function to update 'Teams Goals' and 'Outcome'
    update_teams_goals_and_outcome(df)
    # Define the updated file path
    updated_file_path = os.path.join(directory_paths, f"updated_{csv_file.replace('players', 'updated_players')}")
    # Save the updated DataFrame to a new CSV file without headers
    df.to_csv(updated_file_path, sep=';', index=False, header=True)

  
# #PILES ALL CONVERTED FIELD FILES TOGETHER IN ONE FILE called 'players_rl_data.xlsx'
folder = '/Users/jonco11ins/Documents/players_analytics'
# Define the destination folder where the Excel file will be saved
destination = '/Users/jonco11ins/documents/players_analytics'
delimiter = ';'
writer = pd.ExcelWriter(f'{destination}/players_rl_data.xlsx', engine='openpyxl')
pd.DataFrame().to_excel(writer, sheet_name='Sheet1')
# Iterate through all files in the folder
for file in os.listdir(folder):
    # Check if the file is a CSV file
    if file.endswith('updated_players.csv'):
        # Read the CSV file using the specified delimiter
        df = pd.read_csv(os.path.join(folder, file), delimiter=delimiter)
        # Get the file name without the extension
        file_name = os.path.splitext(file)#[0]
        # Append the DataFrame to the Excel worksheet
        df.to_excel(writer, sheet_name='Sheet1',index=False, header=False, startrow=writer.sheets['Sheet1'].max_row)
# Save the Excel file
writer.save()  


#ADDIND DATES TO FILE
file_path = '/Users/jonco11ins/Documents/players_analytics/players_rl_data.xlsx'
# Read the Excel file
df = pd.read_excel(file_path)
# Get today's date
today = datetime.now().strftime('%Y-%m-%d')
# Update the column 'BW' with today's date
df.loc[:, 'BW'] = today
# Create a blank DataFrame with one row filled with None
blank_index = [0]  # Choose the index where the blank row will be inserted
blank_df = pd.DataFrame(index=blank_index, columns=df.columns)
# Concatenate the blank DataFrame and the original DataFrame
df = pd.concat([blank_df, df], ignore_index=True)
# Save the DataFrame back to the Excel file
df.to_excel(file_path, index=False,header=False)


#Moves all files ending with 'players.csv' to a processed folder to ensure they are not duplicated with false dates
source_dir = '/Users/jonco11ins/documents/players_analytics'
destination_dir = '/Users/jonco11ins/documents/players_analytics/processed'
# Get a list of all files in the source directory
files = os.listdir(source_dir)
# Move all files ending in 'players.csv' to the destination directory
for file in files:
    if file.endswith('players.csv'):
        os.rename(os.path.join(source_dir, file), os.path.join(destination_dir, file))
print('Files moved successfully.')


#ADDS HEADER, REMOVES COLUMN 2, FREEZES ROW 1 DATA IS READY FOR LOADING 
header_path = '/Users/jonco11ins/Documents/players_analytics/header.xlsx'
# Path to the file you want to modify
data_path = '/Users/jonco11ins/Documents/players_analytics/players_rl_data.xlsx'
# Path to the new file
new_path = '/Users/jonco11ins/Documents/players_analytics/RL_DATA.xlsx'
# Read the first row (header) from the header file
header_df = pd.read_excel(header_path, header=None, nrows=1)
# Read the content of the players_rl_data file
data_df = pd.read_excel(data_path, header=None)
# Combine the header with the data
combined_df = pd.concat([header_df, data_df], ignore_index=True)
# Add the 'outcome#' column based on the condition
combined_df['Outcome#'] = combined_df[72].apply(lambda x: 0 if x == 'L' else (1 if x == 'W' else 'Outcome#'))
# Delete the second row (index 1, since indexing starts from 0)
combined_df.drop(index=1, inplace=True)
columns_to_remove = [1,0, 3, 4, 5, 6]
combined_df.drop(columns=columns_to_remove, inplace=True)
# Write the combined DataFrame to the 'RL_DATA.xlsx' file, including the header
combined_df.to_excel(new_path, index=False, header=False)
# Load the workbook using openpyxl
workbook = load_workbook(new_path)
worksheet = workbook.active
# Bold the text in the header row
for cell in worksheet["1:1"]:
    cell.font = Font(bold=True)
# Freeze the top row
worksheet.freeze_panes = "A2"
# Save the modified workbook
workbook.save(new_path)


#Combind new data with what already exists
new_path = '/Users/jonco11ins/Documents/players_analytics/RL_DATA.xlsx'
file2 = '/Users/jonco11ins/Documents/prod_data/RL_DATA_FINAL.xlsx'
# Read the files into DataFrames
df1 = pd.read_excel(new_path)
df2 = pd.read_excel(file2)
# Concatenate the DataFrames
combined_df = pd.concat([df1, df2], ignore_index=True)
# Write the combined DataFrame to a new file
combined_df.to_excel(file2, index=False)


#Delete newly made raw data as it has just been added
if os.path.exists(new_path):
    os.remove(new_path)
    print("File deleted successfully, good job your done")
else:
    print("File not found")